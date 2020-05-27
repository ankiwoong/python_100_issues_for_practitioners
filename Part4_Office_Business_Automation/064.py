import openpyxl
from openpyxl.drawing.image import Image
import os

# 예제 030에서 저장한 박스폴롯 이미지 파일을 가져와서 Image 객체 생성
cwd = os.getcwd()
filename = "boxplot.png"
filepath = os.path.join(cwd, "output", filename)
img = Image(filepath)
print(img)

wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0, title='Image')
wb.remove(wb['Sheet'])

ws['A1'] = "Box Plot"
ws.add_image(img, 'B3')

wb.save(os.path.join(os.getcwd(), "output", "insert_image.xlsx"))
