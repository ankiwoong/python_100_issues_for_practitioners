import requests
from bs4 import BeautifulSoup
import urllib
import openpyxl
import os


# 예제 011의 구글 뉴스 클리핑 함수를 다시 사용
def google_news_clipping_keyword(keyword_input='파이썬', limit=5):

    keyword = urllib.parse.quote(keyword_input)

    base_url = "https://news.google.com"
    url = base_url + "/search?q=" + keyword + "&hl=ko&gl=KR&ceid=KR%3Ako"

    resp = requests.get(url)
    html_src = resp.text
    soup = BeautifulSoup(html_src, 'lxml')

    news_items = soup.select('div[class="xrnccd"]')

    links = []
    titles = []
    contents = []
    agencies = []
    reporting_dates = []
    reporting_times = []

    for item in news_items[:limit]:
        link = item.find('a', attrs={'class': 'VDXfz'}).get('href')
        news_link = base_url + link[1:]
        links.append(news_link)

        news_title = item.find('a', attrs={'class': 'DY5T1d'}).getText()
        titles.append(news_title)

        news_content = item.find('span', attrs={'class': 'xBbh9'}).text
        contents.append(news_content)

        news_agency = item.find(
            'a', attrs={'class': 'wEwyrc AVN2gc uQIVzc Sksgp'}).text
        agencies.append(news_agency)

        news_reporting = item.find(
            'time', attrs={'class': 'WW6dff uQIVzc Sksgp'})
        news_reporting_datetime = news_reporting.get('datetime').split('T')
        news_reporting_date = news_reporting_datetime[0]
        news_reporting_time = news_reporting_datetime[1][:-1]
        reporting_dates.append(news_reporting_date)
        reporting_times.append(news_reporting_time)

    result = {'link': links, 'title': titles, 'contents': contents, 'agency': agencies,
              'date': reporting_dates, 'time': reporting_times}

    return result


if __name__ == "__main__":
    # 함수를 실행하여 뉴스 목록 정리
    search_word = input("검색어를 입력하세요: ")
    news = google_news_clipping_keyword(search_word, 5)

    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title=search_word)
    wb.remove(wb['Sheet'])

    col_nums = {}
    for i, k in enumerate(news.keys()):
        col_nums[k] = i + 1

    for k in col_nums:
        for row_num in range(1, len(news[k])+1):
            if row_num == 1:
                ws.cell(row=row_num, column=col_nums[k]).value = k
            else:
                ws.cell(row=row_num,
                        column=col_nums[k]).value = news[k][row_num-1]

    wb.save(os.path.join(os.getcwd(), "output", "google_news_to_excel.xlsx"))
